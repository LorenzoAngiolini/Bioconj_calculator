import pandas as pd
import sympy as sp
import datetime
from openpyxl import load_workbook

file_lys_dataset = "lys_dataset_calculator.xlsx"
file_lys_smiles = "SMILES_Lys.xlsx"
file_cys_dataset = "cys_dataset_calculator.xlsx"
file_cys_smiles = "SMILES_Cys.xlsx"
workbook = load_workbook(filename=file_lys_dataset)
workbook_lys_smiles = load_workbook(filename=file_lys_smiles)
workbook_2 = load_workbook(filename=file_cys_dataset)
workbook_cys_smiles = load_workbook(filename=file_cys_smiles)
sheet = workbook.active
sheet_lys_smiles = workbook_lys_smiles.active
sheet_2 = workbook_2.active
sheet_cys_smiles = workbook_cys_smiles.active
last_row = sheet.max_row
last_row_lys = sheet_lys_smiles.max_row
last_row_2 = sheet_2.max_row
last_row_cys = sheet_cys_smiles.max_row

def bio_lys_calculator():
    # defining bioconjugation conditions
    print('''Remember, it is not necessary to type the unit of measurement.
Select one of the following mAb:
trx (trastuzumab)
ctx (cetuximab)
4E1REC
cd115
c-0302B17
J08''')
    mAb = input("> ").upper()
    try:
        conc_mAb = float(input("conc_mAb (mg/mL): "))
    except ValueError:
        print("invalid input")
        conc_mAb = float(input("conc_mAb (mg/mL): "))
    mw_mAb = float(input("MW_mAb: "))
    microL_mAb = float(input("𝜇L_mAb: "))
    buffer = input("Buffer (without pH): ").upper()
    pH = float(input("pH: "))
    payload = input("ID_payload: ").upper()
    smiles = input("SMILES: ")
    mw_payload = float(input("MW_payload: "))
    mg_payload = float(input("mg_payload: "))
    eq_payload = float(input("eq_payload: "))
    conc_DMSO = float(input("conc_DMSO (from 0.0 to 1.0): "))
    mg_mAb = ((conc_mAb * microL_mAb) / 1000)
    mmol_mAb = (mg_mAb / mw_mAb)
    mM_payload = 10
    microL_DMSO = ((((mg_payload / mw_payload) * 1000) / mM_payload) * 1000)
    mmol_payload = (mmol_mAb * eq_payload)
    microL_payload = (((mmol_payload * 1000) / mM_payload) * 1000)
    V_gly = 1
    mM_gly = 100
    mw_gly = 75.07
    mmol_gly = ((mM_gly / 1000) * V_gly)
    microL_gly = ((((mmol_mAb * (2.0 * eq_payload)) * 1000) / mM_gly) * 1000)
    mg_gly = (mmol_gly * mw_gly)
    x = sp.symbols('x')
    V_i = microL_mAb + microL_payload + microL_gly
    eq = sp.Eq(conc_DMSO, (x / (V_i + x)))
    solution = sp.solve(eq)
    print('''
To prepare the 10 mM solution of payload''', round(microL_DMSO, 1), '''𝜇L of DMSO are necessary.
To''', (round(microL_mAb, 1)), '''𝜇L of mAb add:''',
round(microL_payload, 1), '''𝜇L of the Payload solution previously activated with NHS.
''', round(microL_gly, 1), '''𝜇L of the 100 mM glycine solution are necessary to quench the conjugation.
Add''', round(solution[0], 2), "𝜇L of DMSO to reach a final concentration of", (conc_DMSO * 100), '''%.
To prepare the 100 mM glycine solution, dissolve ''', round(mg_gly, 1),'''mg in 1 mL of water.''')
    lys_dataset = [mAb, payload, "NHS", buffer, pH, eq_payload]
    lys_smiles = [payload, smiles]
    for col_num, values in enumerate(lys_dataset, start=1):
        sheet.cell(row=last_row + 1, column=col_num, value=values)
    for column_num, smiles_values in enumerate(lys_smiles, start=1):
        sheet_lys_smiles.cell(row=last_row_lys + 1, column=column_num, value=smiles_values)
    data = pd.DataFrame([[conc_mAb, "10mM", "100mM"],
                         [round(mw_mAb, 1), round(mw_payload, 1), round(mw_gly, 1)],
                         ["1", eq_payload, (2*eq_payload)],
                         [round(microL_mAb, 1), round(microL_payload, 1), round(microL_gly, 1)]],
                        index=["Concentration(mg/mL)", "MW", "equivalents", "Volume(𝜇L)"],
                        columns=["mAb", "payload", "Gly"]
                        )
    filename_excel_1 = payload + "_LYS_" + datetime.datetime.now().strftime("%Y-%m-%d") + ".xlsx"
    workbook.save(filename=file_lys_dataset)
    workbook_lys_smiles.save(filename=file_lys_smiles)
    data.to_excel(filename_excel_1)
    return (filename_excel_1)


def in_situ():
    # defining bioconjugation conditions
    print('''Remember, it is not necessary to type the unit of measurement.
Select one of the following mAb:
trx (trastuzumab)
ctx (cetuximab)
4E1REC
cd115
c-0302B17
J08''')
    mAb = input("> ").upper()
    conc_mAb = float(input("conc_mAb (mg/mL): "))
    mw_mAb = float(input("MW_mAb: "))
    microL_mAb = float(input("𝜇L_mAb: "))
    buffer = input("Buffer (without pH): ").upper()
    pH = float(input("pH: "))
    payload = input("ID_payload: ").upper()
    smiles = input("SMILES: ")
    mw_payload = float(input("MW_payload: "))
    mg_payload = float(input("mg_payload: "))
    eq_payload = float(input("eq_payload: "))
    conc_DMSO = float(input("conc_DMSO (from 0.0 to 1.0): "))
    # calculating the parameters
    mg_mAb = ((conc_mAb * microL_mAb) / 1000)
    mmol_mAb = (mg_mAb / mw_mAb)
    mM_payload = 10
    microL_DMSO = ((((mg_payload / mw_payload) * 1000) / mM_payload) * 1000)
    mmol_payload = (mmol_mAb * eq_payload)
    microL_payload = (((mmol_payload * 1000) / mM_payload) * 1000)
    V_gly = 1
    mM_gly = 100
    mw_gly = 75.07
    mmol_gly = ((mM_gly / 1000) * V_gly)
    microL_gly = ((((mmol_mAb * (2.0 * eq_payload)) * 1000) / mM_gly) * 1000)
    mg_gly = (mmol_gly * mw_gly)
    # calculating sulfo-NHS and EDC-HCl
    V_NHS = 1
    V_EDC = 1
    mw_NHS = 217.13
    mw_EDC = 191.70
    mM_NHS = 100
    mM_EDC = 100
    eq_NHS = 2.0
    eq_EDC = 2.0
    mmol_NHS = ((mM_NHS / 1000) * V_NHS)
    mmol_EDC = ((mM_EDC / 1000) * V_EDC)
    microL_NHS = ((((mmol_mAb * (eq_NHS * eq_payload)) * 1000) / mM_NHS) * 1000)
    microL_EDC = ((((mmol_mAb * (eq_EDC * eq_payload)) * 1000) / mM_EDC) * 1000)
    mg_NHS = (mmol_NHS * mw_NHS)
    mg_EDC = (mmol_EDC * mw_EDC)
    x = sp.symbols('x')
    V_i = microL_mAb + microL_payload + microL_gly
    eq = sp.Eq(conc_DMSO, (x / (V_i + x)))
    solution = sp.solve(eq)
    print('''
To prepare the 10 mM solution of payload''', round(microL_DMSO, 1), '''𝜇L of DMSO are necessary.
To''', (round(microL_mAb, 1)), '''𝜇L of mAb add:
''', round(microL_payload, 1), '''𝜇L of Payload in situ activated using:
''', round(microL_NHS, 1), '''𝜇L of sulfo-NHS.
''', round(microL_EDC, 1), '''𝜇L of EDC-HCl.
''', round(microL_gly, 1), '''𝜇L of the 100 mM glycine solution are necessary to quench the conjugation.
Add''', round(solution[0], 2), "𝜇L of DMSO to reach a final concentration of", (conc_DMSO * 100), '''%.
To prepare the 100 mM sulfo-NHS solution, dissolve''', round(mg_NHS, 1), '''mg in 1 mL of water.
To prepare the 100 mM EDC-HCl solution, dissolve''', round(mg_EDC, 1), '''mg in 1 mL of water.
To prepare the 100 mM glycine solution, dissolve ''', round(mg_gly, 1),'''mg in 1 mL of water.''')
    lys_dataset = [mAb, payload, "NHS", buffer, pH, eq_payload]
    lys_smiles = [payload, smiles]
    for col_num, values in enumerate(lys_dataset, start=1):
        sheet.cell(row=last_row + 1, column=col_num, value=values)
    for column_num, smiles_values in enumerate(lys_smiles, start=1):
        sheet_lys_smiles.cell(row=last_row_lys + 1, column=column_num, value=smiles_values)
    data = pd.DataFrame([[conc_mAb, "10mM", "100mM", "100mM", "100mM"],
                         [round(mw_mAb, 1), round(mw_payload, 1), round(mw_NHS, 1), round(mw_EDC), round(mw_gly, 1)],
                         ["1", eq_payload, (2 * eq_payload), (2 * eq_payload), (2 * eq_payload)],
                         [round(microL_mAb, 1), round(microL_payload, 1), round(microL_NHS, 1), round(microL_EDC, 1), round(microL_gly, 1)]],
                        index=["Concentration(mg/mL)", "MW", "equivalents", "Volume(𝜇L)"],
                        columns=["mAb", "payload", "s-NHS", "EDC-HCl", "Gly"]
                        )
    filename_excel_2 = payload + "_LYS_in-situ_" + datetime.datetime.now().strftime("%Y-%m-%d") + ".xlsx"
    workbook.save(filename=file_lys_dataset)
    workbook_lys_smiles.save(filename=file_lys_smiles)
    data.to_excel(filename_excel_2)
    return (filename_excel_2)


def bio_cys_calculator():
    # defining bioconjugation conditions
    print('''Remember, it is not necessary to type the unit of measurement.
    Select one of the following mAb:
    trx (trastuzumab)
    ctx (cetuximab)
    4E1REC
    cd115
    c-0302B17
    J08''')
    mAb = input("> ").upper()
    conj_method = input("Conj_method('onepot' or 'dialyzed'): ").lower()
    conc_mAb = float(input("conc_mAb (mg/mL): "))
    mw_mAb = float(input("MW_mAb: "))
    microL_mAb = float(input("𝜇L_mAb: "))
    reductant = input("reductant: ").upper()
    eq_reductant = float(input("eq_reductant: "))
    mw_reductant = float(input("MW_reductant: "))
    buffer = input("Buffer (without pH): ").upper()
    pH = float(input("pH: "))
    payload = input("ID_payload: ").upper()
    smiles = input("SMILES: ")
    mw_payload = float(input("MW_payload: "))
    mg_payload = float(input("mg_payload: "))
    eq_payload = float(input("eq_payload: "))
    conc_DMSO = float(input("conc_DMSO (from 0.0 to 1.0): "))
    # calculating the parameters
    mg_mAb = ((conc_mAb * microL_mAb) / 1000)
    mmol_mAb = (mg_mAb / mw_mAb)
    V_red = 10
    mM_red = 1
    mmol_red = ((mM_red / 1000) * V_red)
    mg_red = (mmol_red * mw_reductant)
    microL_red = ((((mmol_mAb * eq_reductant) * 1000) / mM_red) * 1000)
    mM_payload = 10
    microL_DMSO = ((((mg_payload / mw_payload) * 1000) / mM_payload) * 1000)
    mmol_payload = (mmol_mAb * eq_payload)
    microL_payload = (((mmol_payload * 1000) / mM_payload) * 1000)
    x = sp.symbols('x')
    V_i = microL_mAb + microL_payload + microL_red
    eq = sp.Eq(conc_DMSO, (x / (V_i + x)))
    solution = sp.solve(eq)
    print('''
To prepare the 10 mM solution of payload''', round(microL_DMSO, 1), '''𝜇L of DMSO are necessary.
To prepare the 10 mM solution of ''', reductant, ',dissolve', round(mg_red, 1), '''mg in 10 mL of water.
To''', (round(microL_mAb, 1)), '''𝜇L of mAb add:
''', round(microL_red, 1), '''𝜇L of the ''', reductant, '''solution.
''', round(microL_payload, 1), '''𝜇L of the Payload solution.
Add''', round(solution[0], 2), "𝜇L of DMSO to reach a final concentration of", (conc_DMSO * 100), '''%.''')
    cys_dataset = [mAb, payload, reductant, eq_reductant, buffer, pH, eq_payload, conj_method]
    for col_num, values in enumerate(cys_dataset, start=1):
        sheet_2.cell(row=last_row_2 + 1, column=col_num, value=values)
    cys_smiles = [payload, smiles]
    for column_num, smiles_values in enumerate(cys_smiles, start=1):
        sheet_cys_smiles.cell(row=last_row_cys + 1, column=column_num, value=smiles_values)
    data = pd.DataFrame([[conc_mAb, "1mM", "10mM"],
                         [round(mw_mAb, 1), round(mw_reductant, 1), round(mw_payload, 1)],
                         ["1", eq_reductant, eq_payload],
                         [round(microL_mAb, 1), round(microL_red, 1), round(microL_payload, 1)]],
                        index=["Concentration(mg/mL)", "MW", "equivalents", "Volume(𝜇L)"],
                        columns=["mAb", "reductant", "payload"]
    )
    filename_excel_3 = payload + "_CYS_" + datetime.datetime.now().strftime("%Y-%m-%d") + ".xlsx"
    workbook_2.save(filename=file_cys_dataset)
    workbook_cys_smiles.save(filename=file_cys_smiles)
    data.to_excel(filename_excel_3)
    return(filename_excel_3)


def help_tool():
    print('''
First of all, you have to explicit the aminoacid of interest for the conjugation.. for example lys (lysine) or cys (cysteine).
Next, the program will ask you to define all the parameters to set the conjugation process. 
Do not specify the unit of measurement, it's not necessary.
To retrieve the SMILES string of your linker-payload: open ChemDraw -> select the molecule of interest -> open the edit window -> CopyAs -> SMILES (shortcut on MacOS: ⌥ + ⌘ + c).
''')