from openpyxl import load_workbook
from openpyxl.styles import Alignment

def excel_modifier(file_path):
    workbook_path = load_workbook(filename=file_path)
    sheet_path = workbook_path.active

    file_lys = "lys_dataset_calculator.xlsx"
    workbook_lys = load_workbook(filename=file_lys)
    sheet_lys = workbook_lys.active

    file_cys = "cys_dataset_calculator.xlsx"
    workbook_cys = load_workbook(filename=file_cys)
    sheet_cys = workbook_cys.active

    sheets = [sheet_path, sheet_lys, sheet_cys]
    widht_columns = 20

    for sheet in sheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='right')
        for column_cells in sheet.columns:
            column_letter = column_cells[0].column_letter
            sheet.column_dimensions[column_letter].width = widht_columns

    workbook_path.save(filename=file_path)
    workbook_lys.save(filename=file_lys)
    workbook_cys.save(filename=file_cys)

    print("Done 😌")
